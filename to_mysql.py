import sys
import pandas as pd
import mysql.connector
from dotenv import load_dotenv
from openpyxl import load_workbook
import datetime
import os

# ── 1. CONNECT ────────────────────────────────────────────────────────────────
load_dotenv()

conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password=os.getenv("DB_PASSWORD"),
    database="tennis_analytics"
)
cursor = conn.cursor()
print("Connected to MySQL.")

# ── 2. CREATE TABLES ──────────────────────────────────────────────────────────
cursor.execute("""
    CREATE TABLE IF NOT EXISTS matches (
        match_id    INT AUTO_INCREMENT PRIMARY KEY,
        player1     VARCHAR(50),
        player2     VARCHAR(50),
        tournament  VARCHAR(100),
        match_date  DATE,
        round       VARCHAR(10),
        surface     VARCHAR(20),
        court       VARCHAR(50)
    );
""")

cursor.execute("""
    CREATE TABLE IF NOT EXISTS points (
        id             INT AUTO_INCREMENT PRIMARY KEY,
        match_id       INT,
        point          INT,
        server         VARCHAR(50),
        returner       VARCHAR(50),
        serve_number   INT,
        serve_dir      VARCHAR(20),
        serve_dir_2nd  VARCHAR(20),
        first_fault    VARCHAR(20),
        rally_length   INT,
        last_shot      VARCHAR(30),
        outcome        VARCHAR(20),
        is_ace         BOOLEAN,
        error_loc      VARCHAR(20),
        point_winner   VARCHAR(50),
        error_maker    VARCHAR(50),
        set1           INT,
        set2           INT,
        gm1            INT,
        gm2            INT,
        point_score    VARCHAR(10),
        gm_winner      INT,
        set_winner     INT,
        return_depth   VARCHAR(10),
        raw_1st        VARCHAR(100),
        raw_2nd        VARCHAR(100),
        FOREIGN KEY (match_id) REFERENCES matches(match_id)
    );
""")
print("Tables created.")

# ── 3. READ MATCH METADATA FROM EXCEL ────────────────────────────────────────
# Usage: python3 to_mysql.py "data/MatchChart_<match>.xlsm" Clay
DEFAULT_FILE = "data/MatchChart 2026 Dove Men+Care Concepcion Felipe Meligeni Alves Juan Pablo Varillas.xlsm"
FILE    = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
surface = sys.argv[2] if len(sys.argv) > 2 else "Clay"

wb = load_workbook(FILE, read_only=True, data_only=True)
ws = wb["MATCH"]

player1    = ws["B1"].value
player2    = ws["B2"].value
date_raw   = ws["B6"].value
tournament = ws["B7"].value
round_val  = ws["B8"].value
court      = ws["B10"].value
surface    = ws["B11"].value

match_date = datetime.datetime.strptime(str(int(date_raw)), "%Y%m%d").date()

print(f"Match: {player1} vs {player2}")
print(f"Tournament: {tournament} | Surface: {surface} | Round: {round_val}")

# ── 4. INSERT MATCH RECORD ────────────────────────────────────────────────────
cursor.execute("""
    INSERT INTO matches (player1, player2, tournament, match_date, round, surface, court)
    VALUES (%s, %s, %s, %s, %s, %s, %s)
""", (player1, player2, tournament, match_date, round_val, surface, court))

match_id = cursor.lastrowid
print(f"Match inserted with ID: {match_id}")

# ── 5. LOAD AND CLEAN PARSED POINTS ──────────────────────────────────────────
df = pd.read_csv("data/match_parsed.csv")

df["is_ace"]       = df["is_ace"].map({True: 1, False: 0}).fillna(0).astype(int)
df["rally_length"] = pd.to_numeric(df["rally_length"], errors="coerce").fillna(0).astype(int)
df["serve_number"] = pd.to_numeric(df["serve_number"], errors="coerce").fillna(0).astype(int)

def clean_row(row):
    return tuple(None if (isinstance(v, float) and str(v) == 'nan') else v for v in row)

# ── 6. INSERT POINTS ──────────────────────────────────────────────────────────
insert_query = """
    INSERT INTO points (
        match_id, point, server, returner, serve_number, serve_dir,
        serve_dir_2nd, first_fault, rally_length, last_shot, outcome,
        is_ace, error_loc, point_winner, error_maker,
        set1, set2, gm1, gm2, point_score, gm_winner, set_winner, return_depth,
        raw_1st, raw_2nd
    ) VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s, %s, %s, %s, %s
    )
"""

rows = []
for row in df.itertuples(index=False):
    cleaned = clean_row(row)
    rows.append((match_id,) + cleaned)

cursor.executemany(insert_query, rows)
conn.commit()

print(f"Inserted {cursor.rowcount} points for match ID {match_id}.")
cursor.close()
conn.close()