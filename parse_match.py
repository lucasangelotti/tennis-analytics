import sys
import pandas as pd
from openpyxl import load_workbook

# ── 1. LOAD THE MATCH FILE ────────────────────────────────────────────────────
DEFAULT_FILE = "data/MatchChart 2026 Dove Men+Care Concepcion Felipe Meligeni Alves Juan Pablo Varillas.xlsm"
FILE = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE

wb = load_workbook(FILE, read_only=True, data_only=True)
ws = wb["MATCH"]

# ── 2. GET PLAYER NAMES ───────────────────────────────────────────────────────
player1 = ws["B1"].value  # serves first
player2 = ws["B2"].value
print(f"Player 1 (serves first): {player1}")
print(f"Player 2: {player2}")

# ── 3. NOTATION LOOKUP TABLES ─────────────────────────────────────────────────
SERVE_DIR  = {"4": "wide", "5": "body", "6": "T", "0": "unknown"}
FAULT_TYPE = {"n": "net", "w": "wide", "d": "deep", "x": "wide_and_deep",
              "g": "foot_fault", "e": "unknown", "!": "shank", "V": "time_violation"}
SHOT_TYPE  = {"f": "forehand", "b": "backhand", "r": "forehand_slice",
              "s": "backhand_slice", "v": "forehand_volley", "z": "backhand_volley",
              "o": "forehand_overhead", "p": "backhand_overhead",
              "u": "unknown_volley", "y": "overhead", "l": "lob",
              "h": "half_volley", "j": "swinging_volley",
              "t": "drop_shot", "q": "drop_shot_backhand"}
ERROR_LOC  = {"n": "net", "w": "wide", "d": "deep", "x": "wide_and_deep", "!": "shank"}
FAULT_SET  = set("nwdxgeV!")

# ── 4. HELPER FUNCTIONS ───────────────────────────────────────────────────────
def get_serve_dir(code):
    if not code:
        return "unknown"
    s = str(code).replace("c", "").strip()
    return SERVE_DIR.get(s[0], "unknown") if s else "unknown"

def get_first_fault(code):
    if not code:
        return None
    s = str(code).replace("c", "").strip()
    if len(s) >= 2 and s[1] in FAULT_SET:
        return FAULT_TYPE.get(s[1], "unknown")
    return None

def get_last_shot(code):
    if not code:
        return None
    s = str(code).replace("c", "").replace("+", "").strip()
    body = s[1:] if s else ""
    body = body.rstrip("*#@nwdx!")
    last = None
    for ch in body:
        if ch in SHOT_TYPE:
            last = SHOT_TYPE[ch]
    return last

def get_error_loc(code):
    if not code:
        return None
    for ch in reversed(str(code)):
        if ch in ERROR_LOC:
            return ERROR_LOC[ch]
    return None

def determine_outcome(is_ace, is_unret, is_rally_winner, is_forced, is_unforced, is_double):
    if is_ace:       return "ace"
    if is_double:    return "double_fault"
    if is_unret:     return "unreturnable"
    if is_rally_winner: return "winner"
    if is_forced:    return "forced_error"
    if is_unforced:  return "unforced_error"
    return "unknown"

def get_return_depth(code):
    """Extract return depth (7/8/9) from rally code — third char of first shot after serve."""
    if not code:
        return None
    s = str(code).replace("c", "").replace("+", "").strip()
    # s[0] = serve direction, s[1] = return shot type, s[2] = return direction or depth
    # depth digit is 7, 8, or 9
    if len(s) >= 4 and s[1] in SHOT_TYPE:
        # check positions 2 and 3 for depth digit
        for i in [2, 3]:
            if i < len(s) and s[i] in ('7', '8', '9'):
                depth_map = {'7': 'short', '8': 'mid', '9': 'deep'}
                return depth_map[s[i]]
    return None

# ── 5. READ AND PARSE ALL POINTS ──────────────────────────────────────────────
parsed = []
point_num = 0

for row in ws.iter_rows(min_row=18, max_row=600, values_only=True):
    if row[0] is None:
        break

    point_num += 1
    server_num      = row[10]
    first_raw       = row[13]
    second_raw      = row[14]
    is_ace          = bool(row[29])
    is_unret        = bool(row[30])
    is_rally_winner = bool(row[31])
    is_forced       = bool(row[32])
    is_unforced     = bool(row[33])
    is_double       = bool(row[34])
    raw_rally_len   = int(row[38]) if row[38] is not None else 0
    rally_len = raw_rally_len
    is_svr_winner   = row[40]
    set1        = row[1]
    set2        = row[2]
    gm1         = row[3]
    gm2         = row[4]
    point_score = str(row[5]) if row[5] is not None else None
    gm_winner   = int(row[42]) if row[42] is not None else 0
    set_winner  = int(row[45]) if row[45] is not None else 0

    if not first_raw and not second_raw:
        continue

    if isinstance(server_num, (int, float)) and int(server_num) == 1:
        server, returner = player1, player2
    elif isinstance(server_num, (int, float)) and int(server_num) == 2:
        server, returner = player2, player1
    else:
        server, returner = "unknown", "unknown"

    first_fault   = get_first_fault(first_raw)
    first_in      = first_fault is None
    serve_num     = 1 if first_in else 2
    serve_dir     = get_serve_dir(first_raw)
    serve_dir_2nd = get_serve_dir(second_raw) if not first_in and second_raw else None
    rally_code    = first_raw if first_in else second_raw
    return_depth = get_return_depth(rally_code)
    last_shot     = get_last_shot(rally_code)
    error_loc     = get_error_loc(rally_code)

    # Point winner read directly from Excel — no heuristics
    if isinstance(is_svr_winner, (int, float)):
        point_winner = server if int(is_svr_winner) == 1 else returner
    else:
        point_winner = None

    outcome = determine_outcome(is_ace, is_unret, is_rally_winner,
                                is_forced, is_unforced, is_double)

    parsed.append({
        "point": point_num, "server": server, "returner": returner,
        "serve_number": serve_num, "serve_dir": serve_dir,
        "serve_dir_2nd": serve_dir_2nd, "first_fault": first_fault,
        "rally_length": rally_len, "last_shot": last_shot,
        "outcome": outcome, "is_ace": is_ace, "error_loc": error_loc,
        "point_winner": point_winner, "error_maker": returner if isinstance(is_svr_winner, (int, float)) and int(is_svr_winner) == 1 else server,
        "set1":        set1,
        "set2":        set2,
        "gm1":         gm1,
        "gm2":         gm2,
        "point_score": point_score,
        "gm_winner":   gm_winner,
        "set_winner":  set_winner,
        "return_depth": return_depth,
        "raw_1st": first_raw, "raw_2nd": second_raw,
    })

# ── 6. BUILD DATAFRAME ────────────────────────────────────────────────────────
cols = ["point", "server", "returner", "serve_number", "serve_dir",
        "serve_dir_2nd", "first_fault", "rally_length", "last_shot",
        "outcome", "is_ace", "error_loc", "point_winner", "error_maker",
        "set1", "set2", "gm1", "gm2", "point_score", "gm_winner", "set_winner", "return_depth",
        "raw_1st", "raw_2nd"]

df = pd.DataFrame(parsed)[cols]

# ── 7. PRINT SUMMARY ──────────────────────────────────────────────────────────
print(f"\nTotal points parsed: {len(df)}")
print("\n--- Sample (first 10 points) ---")
print(df[["point","server","serve_number","serve_dir","rally_length",
          "last_shot","outcome","point_winner"]].head(10).to_string())

print("\n=== MATCH SUMMARY ===")
print(f"\nOutcomes:")
print(df["outcome"].value_counts())
print(f"\nAces: {df['is_ace'].sum()}")
print(f"Winners: {(df['outcome']=='winner').sum()}")
print(f"Unreturnables: {(df['outcome']=='unreturnable').sum()}")
print(f"Unforced errors (incl. double faults): {(df['outcome']=='unforced_error').sum() + (df['outcome']=='double_fault').sum()}")
print(f"  - Rally unforced errors: {(df['outcome']=='unforced_error').sum()}")
print(f"  - Double faults: {(df['outcome']=='double_fault').sum()}")
print(f"Forced errors: {(df['outcome']=='forced_error').sum()}")
print(f"\nServe direction:")
print(df["serve_dir"].value_counts())
print(f"\nAvg rally length: {df['rally_length'].mean():.1f} shots")
print(f"\nPoints won:")
print(df["point_winner"].value_counts())

# ── 8. EXPORT ─────────────────────────────────────────────────────────────────
df.to_csv("data/match_parsed.csv", index=False)
print(f"\nDone! Saved to data/match_parsed.csv")